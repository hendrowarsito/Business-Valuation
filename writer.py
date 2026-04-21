"""
Template Writer
- Load template dengan openpyxl (keep_vba=True, data_only=False)
- Tulis HANYA nilai ke sel input (skip formula)
- Jangan ubah font, style, border, fill, alignment
- Jangan tambah baris/kolom baru
"""
import io, copy
from typing import List, Dict, Any
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import numbers as xl_numbers


def _parse_cell_addr(addr: str):
    """'C8' → (col_idx, row_idx) 1-based"""
    col_str = "".join(c for c in addr if c.isalpha()).upper()
    row_str = "".join(c for c in addr if c.isdigit())
    return column_index_from_string(col_str), int(row_str)


def _format_number(value, unit: str = "Juta") -> float:
    """Normalise value based on display unit."""
    try:
        v = float(str(value).replace(",", "").replace(".", "").strip())
    except (ValueError, TypeError):
        return value
    return v


def write_to_template(template_bytes: bytes, mapping: List[Dict[str, Any]],
                      unit: str = "Juta", log_fn=None) -> bytes:
    """
    Parameters
    ----------
    template_bytes : raw bytes dari file template .xlsx
    mapping        : list dari mapper.map_accounts()
    unit           : "Juta" / "Miliar" / "Penuh"
    log_fn         : callable(msg, level) untuk logging ke Streamlit

    Returns
    -------
    bytes dari workbook yang telah diisi
    """
    wb = load_workbook(io.BytesIO(template_bytes), data_only=False, keep_vba=False)

    written = 0
    skipped_formula = 0
    skipped_null = 0
    errors = []

    for item in mapping:
        tipe = item.get("tipe", "input")
        if tipe in ("formula", "tidak_ditemukan"):
            if tipe == "formula":
                skipped_formula += 1
            continue

        sel = item.get("sel")
        sheet_name = item.get("sheet", wb.sheetnames[0])
        nilai = item.get("nilai")

        if not sel or nilai is None:
            skipped_null += 1
            continue

        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]

        try:
            col_idx, row_idx = _parse_cell_addr(sel)
            cell = ws.cell(row=row_idx, column=col_idx)

            # --- Guard: jangan timpa sel formula ---
            if isinstance(cell.value, str) and cell.value.startswith("="):
                if log_fn:
                    log_fn(f"⚠️  Sel {sel} sudah berisi formula — dilewati", "warn")
                skipped_formula += 1
                continue

            # --- Guard: jangan keluar dari range template ---
            max_row = ws.max_row or 9999
            max_col = ws.max_column or 9999
            if row_idx > max_row or col_idx > max_col:
                errors.append(f"Sel {sel} di luar range template ({max_row}x{max_col})")
                continue

            # --- Tulis nilai (pertahankan format sel) ---
            try:
                numeric_val = float(str(nilai).replace(",", "").replace(" ", ""))
                cell.value = numeric_val
            except (ValueError, TypeError):
                cell.value = nilai

            written += 1

        except Exception as e:
            errors.append(f"Sel {sel}: {e}")

    if log_fn:
        log_fn(f"✅ Sel tertulis: {written}", "ok")
        log_fn(f"∑  Formula dipertahankan: {skipped_formula}", "info")
        if skipped_null:
            log_fn(f"—  Sel dilewati (null): {skipped_null}", "gray")
        for err in errors[:5]:
            log_fn(f"❌ {err}", "err")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def validate_balance_sheet(mapping: List[Dict]) -> List[Dict]:
    """
    Validasi sederhana: Total Aset ≈ Total Liabilitas + Ekuitas
    Returns list of validation messages.
    """
    results = []
    def find(akun_kunci: str):
        for item in mapping:
            if item.get("akun_kunci") == akun_kunci and item.get("nilai") is not None:
                try:
                    return float(str(item["nilai"]).replace(",",""))
                except:
                    return None
        return None

    total_aset = find("total_aset")
    total_liab = find("total_liabilitas")
    total_ekuitas = find("total_ekuitas")

    if all(v is not None for v in [total_aset, total_liab, total_ekuitas]):
        diff = abs(total_aset - (total_liab + total_ekuitas))
        pct = diff / max(total_aset, 1) * 100
        if pct < 0.1:
            results.append({"check": "Balance Sheet", "status": "ok",
                            "msg": f"✅ Total Aset = Total Liabilitas + Ekuitas (selisih {pct:.4f}%)"})
        elif pct < 2:
            results.append({"check": "Balance Sheet", "status": "warn",
                            "msg": f"⚠️  Selisih kecil {pct:.2f}% — mungkin ada pembulatan atau akun yang belum terpetakan"})
        else:
            results.append({"check": "Balance Sheet", "status": "err",
                            "msg": f"❌ Tidak balance: selisih {pct:.1f}% — periksa mapping akun"})
    else:
        results.append({"check": "Balance Sheet", "status": "warn",
                        "msg": "⚠️  Data tidak cukup untuk validasi balance sheet"})

    pendapatan = find("pendapatan")
    hpp = find("hpp")
    laba_kotor = find("laba_kotor")
    if pendapatan and hpp and laba_kotor:
        exp = pendapatan - hpp
        diff = abs(laba_kotor - exp) / max(exp, 1) * 100
        if diff < 1:
            results.append({"check": "Laba Kotor", "status": "ok",
                            "msg": f"✅ Laba Kotor = Pendapatan - HPP (selisih {diff:.2f}%)"})
        else:
            results.append({"check": "Laba Kotor", "status": "warn",
                            "msg": f"⚠️  Laba Kotor selisih {diff:.1f}% dari Pendapatan - HPP"})

    return results


def generate_review_log(mapping: List[Dict], lk_data: dict, proj_data: dict) -> str:
    """Generate a readable review log as plain text."""
    lines = []
    lines.append("=" * 60)
    lines.append("REVIEW LOG — LEMBAR KERJA PENILAIAN SAHAM")
    lines.append("=" * 60)
    lines.append(f"Emiten   : {lk_data.get('nama_emiten','?')}")
    lines.append(f"Kode     : {lk_data.get('kode_saham','?')}")
    lines.append(f"Periode  : {lk_data.get('periode','?')}")
    lines.append(f"Satuan   : {lk_data.get('satuan','?')}")
    lines.append("")

    ok_items = [m for m in mapping if m.get("tipe") == "input" and m.get("confidence", 0) >= 80]
    warn_items = [m for m in mapping if m.get("tipe") == "input" and m.get("confidence", 0) < 80]
    formula_items = [m for m in mapping if m.get("tipe") == "formula"]
    miss_items = [m for m in mapping if m.get("tipe") == "tidak_ditemukan"]

    lines.append(f"RINGKASAN: {len(ok_items)} mapped | {len(warn_items)} perlu review | "
                 f"{len(formula_items)} formula | {len(miss_items)} tidak ditemukan")
    lines.append("")

    if ok_items:
        lines.append("✅ AKUN BERHASIL DIPETAKAN:")
        for m in ok_items:
            lines.append(f"  {m['sel']:6s} {m['nama_akun_template']:<35s}  = {str(m['nilai']):>15s}  (conf {m['confidence']}%)")

    if warn_items:
        lines.append("\n⚠️  AKUN PERLU REVIEW:")
        for m in warn_items:
            lines.append(f"  {m['sel']:6s} {m['nama_akun_template']:<35s}  conf={m['confidence']}%  — {m.get('catatan','')}")

    if formula_items:
        lines.append("\n∑  SEL FORMULA (TIDAK DIUBAH):")
        for m in formula_items:
            lines.append(f"  {m['sel']:6s} {m['nama_akun_template']:<35s}  {m.get('catatan','')}")

    if miss_items:
        lines.append("\n❌ AKUN TIDAK DITEMUKAN DI TEMPLATE:")
        for m in miss_items:
            lines.append(f"  {m['akun_kunci']:<35s}  {m.get('catatan','')}")

    lines.append("\n" + "=" * 60)
    return "\n".join(lines)
