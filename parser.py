import io, json, base64, re
import pandas as pd
import openpyxl
from openpyxl import load_workbook


# ── XLSX → structured text ────────────────────────────────────────────────────

def xlsx_to_text(file_bytes: bytes, max_chars: int = 14000) -> str:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        result = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            result.append(f"\n=== Sheet: {sname} ===")
            for row in ws.iter_rows(max_row=200):
                row_vals = []
                for cell in row:
                    v = cell.value
                    if v is not None:
                        row_vals.append(str(v))
                if any(v.strip() for v in row_vals):
                    result.append("\t".join(row_vals))
        return "\n".join(result)[:max_chars]
    except Exception as e:
        return f"Error reading XLSX: {e}"


def pdf_to_base64(file_bytes: bytes) -> str:
    return base64.standard_b64encode(file_bytes).decode("utf-8")


# ── Template schema reader ────────────────────────────────────────────────────

def read_template_schema(file_bytes: bytes) -> dict:
    """
    Returns:
        {
          "sheets": [{"name": str, "cells": [{"addr": str, "value": any, "formula": str|None, "row": int, "col": int}]}],
          "raw_text": str   -- for sending to Claude
        }
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    sheets = []
    raw_lines = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        cells = []
        raw_lines.append(f"\n=== Sheet: {sname} ===")
        max_row = min(ws.max_row or 150, 200)
        max_col = min(ws.max_column or 30, 40)

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                addr = cell.coordinate
                val = cell.value
                formula = None

                if val is not None:
                    if isinstance(val, str) and val.startswith("="):
                        formula = val[1:]
                        raw_lines.append(f"{addr}: [FORMULA={val}]")
                    else:
                        raw_lines.append(f"{addr}: \"{val}\"")

                    cells.append({
                        "addr": addr, "row": r, "col": c,
                        "value": None if formula else val,
                        "formula": formula,
                        "is_formula": formula is not None
                    })

        sheets.append({"name": sname, "cells": cells})

    return {
        "sheets": sheets,
        "raw_text": "\n".join(raw_lines)[:10000]
    }


# ── Identify formula cells to protect ────────────────────────────────────────

def get_formula_cells(schema: dict) -> set:
    formula_addrs = set()
    for sheet in schema.get("sheets", []):
        for cell in sheet.get("cells", []):
            if cell.get("is_formula"):
                formula_addrs.add(f"{sheet['name']}!{cell['addr']}")
    return formula_addrs


# ── Detect label rows (akun names) in template ───────────────────────────────

def get_label_map(schema: dict) -> dict:
    """Map: cleaned label text → {sheet, addr, row, col, adjacent_data_cols}"""
    label_map = {}
    for sheet in schema.get("sheets", []):
        sname = sheet["name"]
        cell_dict = {(c["row"], c["col"]): c for c in sheet["cells"]}
        max_col = max((c["col"] for c in sheet["cells"]), default=1)

        for cell in sheet["cells"]:
            if cell.get("is_formula"):
                continue
            val = cell.get("value")
            if not isinstance(val, str) or not val.strip():
                continue
            label = val.strip()
            r, c = cell["row"], cell["col"]
            data_cols = []
            for dc in range(c + 1, min(max_col + 1, c + 15)):
                nb = cell_dict.get((r, dc))
                if nb and not nb.get("is_formula"):
                    data_cols.append(dc)
            label_map[label] = {
                "sheet": sname,
                "addr": cell["addr"],
                "row": r,
                "col": c,
                "data_cols": data_cols
            }
    return label_map


def col_letter(n: int) -> str:
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result
