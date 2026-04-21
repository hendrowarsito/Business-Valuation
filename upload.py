import streamlit as st
from utils.session import go_to, mark_done


def render():
    st.markdown("### Step 1 — Upload Dokumen")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("**📊 Laporan Keuangan**")
        f1 = st.file_uploader(
            "Upload laporan keuangan emiten",
            type=["xlsx", "xls", "pdf"],
            key="fu_lk",
            help="Annual Report / Laporan Keuangan IDX format PDF atau XLSX"
        )
        if f1:
            st.success(f"✅ {f1.name}")
            st.caption(f"Ukuran: {f1.size / 1024:.1f} KB")
            st.session_state.uploaded_files["lk"] = f1.read()
            st.session_state.lk_filename = f1.name
            st.session_state.lk_is_pdf = f1.name.lower().endswith(".pdf")
        elif "lk" in st.session_state.uploaded_files:
            st.info(f"✅ {st.session_state.get('lk_filename','file')} (sudah diupload)")

    with col2:
        st.markdown("**📈 Proyeksi Keuangan**")
        f2 = st.file_uploader(
            "Upload proyeksi keuangan",
            type=["xlsx", "xls"],
            key="fu_proj",
            help="Model proyeksi keuangan 3-5 tahun ke depan"
        )
        if f2:
            st.success(f"✅ {f2.name}")
            st.caption(f"Ukuran: {f2.size / 1024:.1f} KB")
            st.session_state.uploaded_files["proj"] = f2.read()
            st.session_state.proj_filename = f2.name
        elif "proj" in st.session_state.uploaded_files:
            st.info(f"✅ {st.session_state.get('proj_filename','file')} (sudah diupload)")

    with col3:
        st.markdown("**📋 Template Lembar Kerja**")
        f3 = st.file_uploader(
            "Upload format lembar kerja penilaian",
            type=["xlsx", "xls"],
            key="fu_tpl",
            help="Template XLSX yang formatnya harus dijaga"
        )
        if f3:
            st.success(f"✅ {f3.name}")
            st.caption(f"Ukuran: {f3.size / 1024:.1f} KB")
            st.session_state.uploaded_files["template"] = f3.read()
            st.session_state.tpl_filename = f3.name
        elif "template" in st.session_state.uploaded_files:
            st.info(f"✅ {st.session_state.get('tpl_filename','file')} (sudah diupload)")

    st.markdown("---")

    has_all = all(k in st.session_state.uploaded_files for k in ["lk", "proj", "template"])
    count = sum(1 for k in ["lk", "proj", "template"] if k in st.session_state.uploaded_files)

    st.markdown(f"**Status upload:** {count}/3 dokumen siap")

    if not has_all:
        st.info("Upload ketiga dokumen untuk melanjutkan ke analisis AI.")
        
        with st.expander("💡 Tips format dokumen"):
            st.markdown("""
**Laporan Keuangan (LK):**
- Format IDX: Laporan Tahunan / Laporan Keuangan Audited
- Minimal 2-3 tahun historis
- PDF dari IDX.co.id atau XLSX dari sistem internal

**Proyeksi Keuangan:**
- Sheet dengan kolom per tahun proyeksi
- Baris: Pendapatan, EBITDA, Laba Bersih, FCF, CapEx
- Opsional: asumsi WACC dan terminal growth rate

**Template Lembar Kerja:**
- Format XLSX yang sudah berisi label akun
- Sel formula (subtotal/total) akan **otomatis dideteksi dan dijaga**
- Font, style, border **tidak akan diubah**
            """)
        return

    with st.expander("🔍 Pratinjau template", expanded=False):
        try:
            import openpyxl, io
            wb = openpyxl.load_workbook(
                io.BytesIO(st.session_state.uploaded_files["template"]),
                data_only=True
            )
            st.caption(f"Sheet: {', '.join(wb.sheetnames)}")
            ws = wb.active
            rows = []
            for r in ws.iter_rows(max_row=20, values_only=True):
                row = [str(c) if c is not None else "" for c in r]
                if any(v.strip() for v in row):
                    rows.append(row)
            if rows:
                import pandas as pd
                df = pd.DataFrame(rows)
                st.dataframe(df, use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"Tidak bisa preview: {e}")

    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        if st.button("🚀 Mulai Analisis AI →", type="primary", use_container_width=True):
            mark_done("upload")
            go_to("analysis")
            st.rerun()
    with col_info:
        st.caption("AI akan membaca dan mengekstrak data dari ketiga dokumen secara otomatis.")
